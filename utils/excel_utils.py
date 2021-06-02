from utils.logger_utils import get_logger
from configs import config
import openpyxl

log = get_logger()


class ExcelUtils:
    log.info("Implementation of ExcelUtils")

    def read_from_excel():
        log.info("Inside read_from_excel()")
        complete_test_data = []
        wb = openpyxl.load_workbook(config.TESTDATA_PATH)
        sheet = wb.active
        for row in range(2, sheet.max_row + 1):
            test_data = []
            for column in range(1, sheet.max_column):
                test_data.append(sheet.cell(row, column).value)
            complete_test_data.append(test_data)
        log.info("Finished read_from_excel()")
        return complete_test_data
